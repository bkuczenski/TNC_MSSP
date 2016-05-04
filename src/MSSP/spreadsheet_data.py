"""
Created on Oct 28, 2015

@author: brandon
This is a set of functions for reading grid and header content from the 
MSSP spreadsheets, and formatting it into some semblance of a data
structure.

Overall strategy:
the object types are defined based on the spreadsheet content each time
(for later semantic management, perhaps?)


Since there are three different MSSP spreadsheets, this class has at least
three different functions to load the spreadsheets, however the processing
is different for each spreadsheet:
 * Monitoring
   - questions by columns
   - monitoring options by rows
   - annotations in table grid
   
 * Assessment
   - assessment methods by columns
   - indices by rows (OPEN QUESTION: relation of these to Qs)
   - minimum criteria in table grid
   - some annotations also
   
 * Harvest Control Rules
   - questions by rows
   - rules by columns
   - annotations in table grid

From that come the entity types:
 - monitoring options
 - indices
 - questions
 - criteria
    
"""


import openpyxl as xl
from openpyxl.utils import range_boundaries
from MSSP.exceptions import *
from MSSP.records import Question, Target
from MSSP.elements import *
from MSSP import mssp_work
from MSSP.utils import *

import pandas as pd


def load_default_set():
    m = SpreadsheetData()
    m.parse_all_sheets()
    m.load_mappings(mssp_work.QUESTION_MAP)

    return m


class SpreadsheetData(object):
    """
    Tools for extracting and refactoring information contained in MSSP spreadsheets

    MSSP = Management Strategy Selection Process
    Consists of a collection of 3 static spreadsheets that each contain a set of questions
    and a set of targets.

    An Element is the basic unit of spreadsheet content. It consists of a cell's value,
    text color, and background color. Every element within an ElementSet is unique across
    all three of these characteristics. Each element may have multiple references.

    The routines create four ElementSets:
     - QuestionAttributes - text annotations of questions
     - TargetAttributes - text annotations of targets
     - Criteria - table data pertaining to criteria questions
     - Caveats - table data pertaining to caveat questions

    The routines construct

    The MSSP class itself

    """

    def _open_worksheet(self, sel):
        if check_sel(sel):
            print mssp_work.MSSP_FILES[sel]
            x = xl.load_workbook(self.working_dir + self.files[sel])
            if 'Master' in x.get_sheet_names():
                return x['Master']
            else:
                return x.active

    def __init__(self, version='default', workdir=None,
                 files=mssp_work.MSSP_FILES, grid_start=mssp_work.grid_start,
                 answer_senses=mssp_work.answer_senses):
        """
        Constructor.
        Creates an MSSP object which can be used as a base to perform read-in functions.

        Input kwargs:
         version string - maybe to be used later
         workdir - working directory (default '~/Dropbox/YYYY/TNCWebTool/Working')
         files - dictionary maps selector strings to excel files
         grid_start - dictionary maps selector strings to cell references- top-left of data region
         answer_senses - dictionary maps selector strings to record reference containing
         applicable answer (for 'caveat' questions only - required) (record is (None, col) or (row, None))

         files and grid_start are required; valid_answers can be None

         These variables should be defined in the script that calls the constructor

        """
        try:
            version = mssp_work.Version
        finally:
            self.version = version

        self.Attributes = ElementSet()  # these appear in the header regions of the spreadsheets
        self.Notations = ElementSet()   # these appear in the data regions of the spreadsheets

        # Questions and Targets are dicts where key is 2-tuple: (sel, record)
        # where record is either (None, col) or (row, None)
        self.Questions = {}
        self.Targets = {}

        if workdir is None:
            self.working_dir = defaultdir + 'Working/'
        else:
            self.working_dir = workdir

        if files is None:
            raise EmptyInputError("Please supply dictionary of MSSP files.")

        self.files = files

        if grid_start is None:
            raise EmptyInputError("Please supply dictionary of grid starting cells.")

        self.grid_start = grid_start

        if answer_senses is None:
            print("No answer_senses provided; using last attribute row/column for question_sense")
        self.answer_senses = answer_senses

        self.colormap = None

    @staticmethod
    def cell_in_range(row, col, rng):
        """
        :param row: row index
        :param col: column index
        :param rng: a string describing a range
        :return:
        """
        rn_coords = range_boundaries(rng)
        return ((rn_coords[0] <= col <= rn_coords[2]) &
                (rn_coords[1] <= row <= rn_coords[3]))

    @staticmethod
    def _element_or_merged_range(sheet, row, col):
        """
        If ref is contained within a merged range on sheet, returns the element describing the range.
        Otherwise, returns the element referenced.
        :param sheet:
        :param row:
        :param col:
        :return:
        """
        for rn in sheet.merged_cell_ranges:
            if SpreadsheetData.cell_in_range(row, col, rn):
                return Element.from_worksheet(sheet, rn)

        try:
            return Element.from_worksheet(sheet, (row, col))
        except EmptyInputError:
            return Element.empty_element()

    @staticmethod
    def _expand_attribute_refs(start, record):
        """
        generates a list of 2-tuples of attribute cells for a given record
        :param start: the start of the table grid
        :param record: 2-tuple (None, col) or (row, None)
        :return: list of (row,col) tuples
        """
        if record[0] is None:  # columnwise definition
            refs = [(row, record[1]) for row in range(1, start.row)]
        elif record[1] is None:  # row-wise definition
            refs = [(record[0], col) for col in range(1, start.col_idx)]
        else:
            raise MsspError('bad record definition {}'.format(record))
        return refs

    def _attributes_of_record(self, sheet, start, record):
        """
        Returns a list of elements belonging to the attribute range of a given record, including merged
        cells if applicable
        :param sheet: worksheet
        :param start: grid data start
        :param record: (None, col) or (row, None)
        :return: list of unique entries in the Attributes ElementSet
        """
        elts = []
        refs = self._expand_attribute_refs(start, record)
        for row, col in refs:
            elt = self._element_or_merged_range(sheet, row, col)
            if elt.text is not None:
                elts.append(elt)

        inds = self.Attributes.add_elements(elts)
        return [self.Attributes[i] for i in inds]

    @staticmethod
    def _expand_record_refs(sheet, start, q_rows):
        """
        Returns iterable question_records, target_records
        :param sheet:
        :param start:
        :param q_rows: flag indicating questions in rows
        :return:
        """
        row_records = [(row, None) for row in range(start.row, sheet.max_row+1)]
        col_records = [(None, col) for col in range(start.col_idx, sheet.max_column+1)]
        if q_rows is True:
            return row_records, col_records
        else:
            return col_records, row_records

    def _grid_elements(self, sheet, start, record):
        """
        Returns a list of data elements for a given reference, given in record
        :param sheet:
        :param start:
        :param record: record specification (None, col) or (row, None)
        :return: a list of (cross-record, element) tuples
        """
        mapping = []  # a list of cross-record-to-element mappings
        if record[0] is None:
            # column-spec
            for row in range(start.row, sheet.max_row+1):
                try:
                    elt = Element.from_cell(sheet.cell(None, row, record[1]))
                except EmptyInputError:
                    continue
                i = self.Notations.add_element(elt)
                mapping.append(((row, None), self.Notations[i]))
        else:  # row-spec
            for col in range(start.col_idx, sheet.max_column+1):
                try:
                    elt = Element.from_cell(sheet.cell(None, record[0], col))
                except EmptyInputError:
                    continue
                i = self.Notations.add_element(elt)
                mapping.append(((None, col), self.Notations[i]))

        return mapping

    def _get_answer_sense(self, sheet, sel, record):
        """
        Extract the correct answer sense for a given question record from the specified answer_sense reference if
        present, or from the final attribute row/column if not.
        :param sheet: open worksheet
        :param sel: valid selector
        :param record: question for which answer sense is sought
        :return: cell contents containing the answer sense
        """
        if self.answer_senses is None:
            start = sheet[self.grid_start[sel]]  # a cell
            # select last attribute cross-record
            answer_record = (start.row-1, start.col_idx-1)
        else:
            # select specified cross-record
            answer_record = self.answer_senses[sel]

        if record[0] is not None:  # column-wise
            return self._element_or_merged_range(sheet, record[0], answer_record[1]).text
        elif record[1] is not None:  # row-wise
            return self._element_or_merged_range(sheet, answer_record[0], record[1]).text
        else:
            raise MsspError('bad record reference {}'.format(record))

    def parse_file(self, sel, q_rows=True):
        """
        Method to parse the monitoring spreadsheet automagically.
        This function will add new entries to all four ElementSets

        First step: create new questions, each one having only QuestionAttributes
          - here we have to deal with the merging
        Second step: create new targets
          - here we have to deal with the merging
        Third step: read in Criteria from Criteria questions; associate
        Fourth step: read in

        :param sel:
        :param q_rows: whether questions are in rows (default) or columns
        :return:
        """
        if not check_sel(sel):
            raise BadSelectorError

        sheet = self._open_worksheet(sel)
        start = sheet[self.grid_start[sel]]  # a cell
        q_records, t_records = self._expand_record_refs(sheet, start, q_rows)

        print "adding questions"
        for record in q_records:
            mapped_attrs = self._attributes_of_record(sheet, start, record)
            answer_sense = self._get_answer_sense(sheet, sel, record)
            # create a new question with those attributes
            self.Questions[(sel, record)] = Question(sel, record, mapped_attrs, answer_sense=answer_sense)

        print "adding targets"
        for record in t_records:
            mapped_attrs = self._attributes_of_record(sheet, start, record)
            # create a new target with those attributes
            self.Targets[(sel, record)] = Target(sel, record, mapped_attrs)

        print "populating questions"
        for record in q_records:
            try:
                q = self.Questions[(sel, record)]
            except KeyError:
                continue
            if q.criterion is True:
                # index all grid elements, add to Criteria ElementSet;
                mappings = self._grid_elements(sheet, start, record)
                q.encode_criteria(mappings)
            else:
                mappings = self._grid_elements(sheet, start, record)
                q.encode_caveats(mappings)

            for mapping in mappings:
                try:
                    t = self.Targets[(sel, mapping[0])]
                except KeyError:
                    continue
                if q.criterion is True:
                    t.add_criteria_mapping((record, mapping[1]))
                else:
                    t.add_caveat_mapping((record, mapping[1]))

        return True

    def parse_monitoring_sheet(self):
        """
        :return:
        """
        return self.parse_file('Monitoring', q_rows=False)

    def parse_assessment_sheet(self):
        return self.parse_file('Assessment', q_rows=True)

    def parse_controlrules_sheet(self):
        return self.parse_file('ControlRules', q_rows=True)

    def parse_all_sheets(self):
        self.parse_monitoring_sheet()
        self.parse_assessment_sheet()
        self.parse_controlrules_sheet()

    def load_mappings(self, qmap_file):
        self.colormap = pd.read_excel(self.working_dir + qmap_file, 'COLORMAP')

        qmap = pd.read_excel(self.working_dir + qmap_file, self.version)
        # now load synonym and satisfies relations into Question objects

        for i,v in qmap.iterrows():
            subj = convert_subject_to_reference(v['Subject'])
            obj = convert_subject_to_reference(v['Object'])
            if v['Relation'] == 'synonym':
                synonyms = set([subj, obj])
                synonyms = synonyms.union(self.Questions[subj].synonyms)
                synonyms = synonyms.union(self.Questions[obj].synonyms)
                for k in synonyms:
                    self.Questions[k].add_synonyms(synonyms)

            if v['Relation'] == 'satisfies':
                self.Questions[obj].add_satisfied(subj)


