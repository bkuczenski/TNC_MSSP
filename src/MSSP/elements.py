"""
Created on Oct 28, 2015

@author: brandon

Entity classes defined here

Element: any text appearing in the spreadsheet. abstract.
to be continued.  Properties include text string, text color, background color (both of which should be RGBs)

Element has __hash__ and __eq__

ElementSet: a collection of elements with a uniqueness condition.  No two Elements in an ElementSet may be
equal in all three fields.

Each MSSP class instance contains two different ElementSets: Attributes (derived from spreadsheet header
content) and Flags (derived from spreadsheet data content).  Possibly these should be per-spreadsheet, but at the
moment I can't think of any good reason why they would need to be.

The first step in populating an MSSP instance is to generate its ElementSets.

OK- attribute is just a list() that overrides append() to check uniqueness

Item: contains a spreadsheet reference ( sheet, row, column )
where either row or column is None
and a list of Attributes

Question: is an Item with an Answer Set

Target: is an Item with a TargetType (Monitoring, Assessment, ControlRule)

The second step in populating an MSSP instance is to generate a



 * Questions
   - have user prompts
   - have valid answer sets
   
 * Monitoring

suggestion from Owen: use "factories" instead of a constructor

class Element(object):
    @classmethod
    def cell(cls, cell):
        ... checks from line 89 ...
        text = cell.value
        text_color = getattr(cell.font.color, cell.font.color.type)
        text_color_type = cell.font.color.type
        ... more attributes ...
        return cls(text, text_color, text_color_type, ...more attributes...)

    def __init__(self, text, text_color, text_color_type, ...)
        self.text = text
        self.text_color = text_color
        self.text_color_type = text_color_type

"""

#from openpyxl.styles.colors import COLOR_INDEX
from MSSP.exceptions import *
import re


class Element(object):
    """
    A decorated text string that appears in the spreadsheet
    """

    def __init__(self, text, text_color=0, fill_color='00000000', ref=None):
        self.text = text
        self.text_color = text_color
        self.fill_color = fill_color
        self.ref = ref

    @staticmethod
    def from_worksheet(worksheet, ref):
        if worksheet is None or ref is None:
            raise MsspError("Both worksheet+ref arguments required")

        if isinstance(ref, str):
            if ':' in ref:
                ref = re.split(':', ref)[0]
            cell = worksheet[ref]
        else:
            cell = worksheet.cell(None, ref[0], ref[1])

        return Element.from_cell(cell)

    @classmethod
    def from_cell(cls, cell):
        """
        Element() constructor accepts an openpyxl.cell.cell.Cell object and returns
        an Element.

        Use .from_worksheet(worksheet,ref) to create an Element by reference.

        :param cell: an openpyxl.cell.cell.Cell object
        :return: an Element
        """
        if isinstance(cell, Element):
            return cell

        if cell is None:
            raise MsspError("No cell input found")

        else:
            if cell.value == 'N/A':
                cell.value = None

            if cell.value is None:
                raise EmptyInputError

            text = cell.value
            text_color = getattr(cell.font.color, cell.font.color.type)
            # TODO: convert indexed colors to RGB-
            # probably in a special function in an openpyxl interfacing module

            # self.text_color_type = cell.font.color.type  # don't need this

            fill_color = getattr(cell.fill.fgColor, cell.fill.fgColor.type)
            # TODO: convert indexed colors to RGB-
            # probably in a special function in an openpyxl interfacing module
            # also TODO: convert theme colors to RGB-
            # no help here; requires xml
            ##self.color_type = cell.fill.fgColor.type

            ref = cell.parent.title + '!' + cell.column + str(cell.row)
            return cls(text, text_color, fill_color, ref)

    def __hash__(self):
        return hash((self.text, self.text_color, self.fill_color))

    def __eq__(self, other):
        if other is None:
            return False
        return ((self.text == other.text) &
                # (self.text_color == other.text_color) &  # compare on text color -> false negatives (libreoffice)
                (self.fill_color == other.fill_color))

    def __ne__(self, other):
        return not self.__eq__(other)

    def __str__(self):
        try:
            return "{0}: {1} [text {2} | fill {3}]".format(self.ref, self.text, self.text_color, self.fill_color)
        except UnicodeEncodeError:
            return "unicode problem with {0}".format(self.ref)

    def search(self, string):
        """
        returns true if string (regex) is found in self.text
        :param string: regex
        :return: bool
        """
        return bool(re.match(string, unicode(self.text), flags=re.IGNORECASE | re.U))


class ElementSet(object):
    """
    An ordered set of unique elements.
    Elements have __hash__ and __eq__ defined, which allows them to be used
    in dicts. This class accepts an element as input and returns its index in
    the ordered set, appending if it does not exist and doing a lookup if it
    does.
    Returns the index into the ElementSet of the original element
    """

    def _append(self, element):
        if element is None:
            return

        if element in self.index:
            k = self.index[element]
            if element.ref not in self.refs[k]:
                self.refs[k].append(element.ref)
        else:
            self.index[element] = len(self.elements)  # the element's index
            self.elements.append(element)    # the element
            self.refs.append([element.ref])  # a list of cell references

        return self.index[element]

    #def add(self, *args, **kwargs):
    #    try:
    #        element = Element(*args, **kwargs)
    #    except EmptyInputError:
    #        # quit without errors if cell has no text
    #        return
    #
    #    return self._append(element)

    def add_element(self, element):
        """
        :param element:
        :return: the index into the ElementSet corresponding to the added element
        """
        return self._append(element)

    def add_elements(self, elements):
        """
        :param elements: list of elements or Nones
        :return: list of unique indices into the ElementSet corresponding to input list
        """
        seq = []
        for element in elements:
            if element is not None:
                k = self._append(element)
                seq.append(k)

        return list(set(seq))  # uniquify seq

    def __getitem__(self, item):
        """
        Lookup item by index
        :param item:
        :return:
        """
        return self.elements[item]

    def get_index(self, elt):
        """
        Lookup index by item
        :param elt:
        :return:
        """
        if isinstance(elt, Element):
            return self.index[elt]
        else:
            return self.index[Element(elt)]

    def __iter__(self):
        return iter(self.elements)

    def __len__(self):
        return len(self.elements)

    def __init__(self):
        self.index = dict()
        self.elements = list()
        self.refs = list()

    def __str__(self):
        st = ""
        for i in self.elements:
            st += str(i) + "\n"
        return st

    def search(self, string, idx):
        """
        Search the element set for a plain string or regex. returns a list of indices
        :param string:
        :return:
        """
        ind = []
        type(string)
        for elt in self.elements:
            if elt.search(string) is True:
                ind.append(self.index[elt])

        if idx is None:
            return ind

        return list(set(ind).intersection(idx))


def row_col_from_cell(cell):
    """

    :param cell: a cell string e.g. 'B12'
    :return: row, col in numerical index
    """
    from openpyxl.utils import column_index_from_string, coordinate_from_string
    (col,row) = coordinate_from_string(cell)
    return row, column_index_from_string(col)


def compare_sheets(file1, sheet1, file2, sheet2, start='A1'):
    import openpyxl as xl
    x1 = xl.load_workbook(file1)
    s1 = x1[sheet1]
    x2 = xl.load_workbook(file2)
    s2 = x2[sheet2]

    if s1.max_row != s2.max_row:
        print "Sheets have different numbers of rows: S1: {0} S2: {1}".format(s1.max_row, s2.max_row)
    if s1.max_column != s2.max_column:
        print "Sheets have different numbers of columns: S1: {0} S2: {1}".format(s1.max_column, s2.max_column)

    min_row, min_col = row_col_from_cell(start)

    max_row = min(s1.max_row, s2.max_row)
    max_col = min(s1.max_column, s2.max_column)

    print "Comparing from [{0},{1}] to [{2},{3}]...".format(min_row, min_col, max_row, max_col)

    dif_list1 = []
    dif_list2 = []

    for row in range(min_row-1, max_row):
        for col in range(min_col-1, max_col):
            try:
                E1 = Element.from_cell(s1.cell(None, row+1, col+1))
            except EmptyInputError:
                E1 = None
            try:
                E2 = Element.from_cell(s2.cell(None, row+1, col+1))
            except EmptyInputError:
                E2 = None

            if E1 != E2:
                dif_list1.append(E1)
                dif_list2.append(E2)

    print "{0} differences found.".format(len(dif_list1))

    if len(dif_list1) < 50:
        for i in range(0,len(dif_list1)):
            print "{0}\n{1}\n---".format(dif_list1[i], dif_list2[i])
    else:
        print "Suppressing long output."

    return dif_list1, dif_list2

